import openpyxl
#read from sheet
data={}
book = openpyxl.load_workbook("//Users//kavin//Desktop//TTc 2.xlsx")
sheet = book.active #it's is used select active sheet in exl
cell = sheet.cell(row=3,column=2)
print(cell.value)

#write
sheet.cell(row=3,column=2).value="lk"
print(sheet.cell(row=3,column=2).value)

#to get number of row and column in xl
print(sheet.max_row)
print(sheet.max_column)

print(sheet["B3"].value)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=="t3":
        for j in range(1,sheet.max_column+1):
            data[sheet.cell(row=i,column=j).value]
            print(sheet.cell(row=i,column=j).value)